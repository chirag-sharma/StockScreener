# services/excelExporter.py
"""
This module provides functionality to export stock analysis results to an Excel file with conditional formatting.
Enhanced with precise value-based color coding for better visual analysis.
"""
import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from stock_screener.core.constants import THRESHOLDS, BASE_OUTPUT_DIR
import os


class ExcelExporter:
    """
    Handles exporting analysis results to an Excel file with enhanced conditional formatting 
    based on precise value thresholds and performance indicators.
    """
    def __init__(self, output_path=None):
        """
        Initialize the exporter with enhanced color coding based on value performance.
        """
        if output_path is None:
            output_path = os.path.join(BASE_OUTPUT_DIR, 'value_analysis.xlsx')
        self.output_path = output_path
        
        # Enhanced color scheme for precise value-based formatting
        self.excellent_fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')      # Dark Green
        self.good_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')          # Light Green  
        self.acceptable_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')    # Yellow
        self.poor_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')          # Orange
        self.bad_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')           # Red
        self.neutral_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')      # Gray
        
        # Font styles
        self.bold_font = Font(bold=True)
        self.bold_white_font = Font(bold=True, color='FFFFFF')
        
        # Define value ranges for each metric type
        self.metric_ranges = self._define_metric_ranges()
    
    def _define_metric_ranges(self):
        """
        Define color coding ranges for each financial metric based on value investing principles.
        Returns a dictionary with metric names and their scoring thresholds.
        """
        return {
            # Valuation Metrics (Lower is Better)
            'PE Ratio': {
                'excellent': (0, 10),      # Dark Green: Undervalued
                'good': (10, 15),          # Light Green: Fair value
                'acceptable': (15, 20),    # Yellow: Slightly overvalued  
                'poor': (20, 25),          # Orange: Overvalued
                'bad': (25, float('inf'))  # Red: Extremely overvalued
            },
            'Price to Book': {
                'excellent': (0, 1),
                'good': (1, 1.5),
                'acceptable': (1.5, 2),
                'poor': (2, 3),
                'bad': (3, float('inf'))
            },
            'EV/EBITDA': {
                'excellent': (0, 6),
                'good': (6, 8),
                'acceptable': (8, 10),
                'poor': (10, 12),
                'bad': (12, float('inf'))
            },
            'Price to Cash Flow': {
                'excellent': (0, 8),
                'good': (8, 12),
                'acceptable': (12, 15),
                'poor': (15, 20),
                'bad': (20, float('inf'))
            },
            
            # Profitability Metrics (Higher is Better)
            'ROE': {
                'excellent': (25, float('inf')),
                'good': (20, 25),
                'acceptable': (15, 20),
                'poor': (10, 15),
                'bad': (0, 10)
            },
            'Return on Assets (ROA)': {
                'excellent': (25, float('inf')),
                'good': (20, 25),
                'acceptable': (15, 20),
                'poor': (10, 15),
                'bad': (0, 10)
            },
            'Net Profit Margin': {
                'excellent': (25, float('inf')),
                'good': (20, 25),
                'acceptable': (15, 20),
                'poor': (10, 15),
                'bad': (0, 10)
            },
            'Operating Margin': {
                'excellent': (25, float('inf')),
                'good': (20, 25),
                'acceptable': (15, 20),
                'poor': (10, 15),
                'bad': (0, 10)
            },
            
            # Financial Strength Metrics (Higher is Better)
            'Current Ratio': {
                'excellent': (2.5, float('inf')),
                'good': (2.0, 2.5),
                'acceptable': (1.5, 2.0),
                'poor': (1.0, 1.5),
                'bad': (0, 1.0)
            },
            'Quick Ratio': {
                'excellent': (2.5, float('inf')),
                'good': (2.0, 2.5),
                'acceptable': (1.5, 2.0),
                'poor': (1.0, 1.5),
                'bad': (0, 1.0)
            },
            'Interest Coverage Ratio': {
                'excellent': (10, float('inf')),
                'good': (7, 10),
                'acceptable': (5, 7),
                'poor': (3, 5),
                'bad': (0, 3)
            },
            
            # Debt Metrics (Lower is Better)
            'Debt/Equity': {
                'excellent': (0, 0.3),
                'good': (0.3, 0.5),
                'acceptable': (0.5, 0.8),
                'poor': (0.8, 1.0),
                'bad': (1.0, float('inf'))
            },
            'Pledged Shares (%)': {
                'excellent': (0, 5),
                'good': (5, 10),
                'acceptable': (10, 20),
                'poor': (20, 30),
                'bad': (30, float('inf'))
            },
            
            # Growth Metrics (Higher is Better)
            'EPS Growth (%)': {
                'excellent': (25, float('inf')),
                'good': (15, 25),
                'acceptable': (10, 15),
                'poor': (5, 10),
                'bad': (0, 5)
            },
            'Revenue Growth (%)': {
                'excellent': (25, float('inf')),
                'good': (15, 25),
                'acceptable': (10, 15),
                'poor': (5, 10),
                'bad': (0, 5)
            },
            
            # Ownership & Quality Metrics
            'Promoter Holding': {
                'excellent': (70, float('inf')),
                'good': (60, 70),
                'acceptable': (50, 60),
                'poor': (40, 50),
                'bad': (0, 40)
            },
            
            # Special Metrics
            'Margin of Safety (%)': {
                'excellent': (30, float('inf')),
                'good': (20, 30),
                'acceptable': (10, 20),
                'poor': (0, 10),
                'bad': (-float('inf'), 0)
            },
            'Cash Conversion Ratio': {
                'excellent': (1.5, float('inf')),
                'good': (1.2, 1.5),
                'acceptable': (1.0, 1.2),
                'poor': (0.8, 1.0),
                'bad': (0, 0.8)
            },
            
            # Value Score (0-100)
            'Value Score': {
                'excellent': (90, 100),
                'good': (70, 90),
                'acceptable': (50, 70),
                'poor': (30, 50),
                'bad': (0, 30)
            },
            'Value Score (1-10)': {
                'excellent': (8, 10),
                'good': (6, 8),
                'acceptable': (4, 6),
                'poor': (2, 4),
                'bad': (0, 2)
            },
            
            # Monthly Growth Predictions (Higher is Better)
            'Growth 6M (%)': {
                'excellent': (15, float('inf')),
                'good': (8, 15),
                'acceptable': (3, 8),
                'poor': (0, 3),
                'bad': (-float('inf'), 0)
            },
            'Growth 7M (%)': {
                'excellent': (15, float('inf')),
                'good': (8, 15),
                'acceptable': (3, 8),
                'poor': (0, 3),
                'bad': (-float('inf'), 0)
            },
            'Growth 8M (%)': {
                'excellent': (15, float('inf')),
                'good': (8, 15),
                'acceptable': (3, 8),
                'poor': (0, 3),
                'bad': (-float('inf'), 0)
            },
            'Growth 9M (%)': {
                'excellent': (15, float('inf')),
                'good': (8, 15),
                'acceptable': (3, 8),
                'poor': (0, 3),
                'bad': (-float('inf'), 0)
            },
            'Growth 10M (%)': {
                'excellent': (15, float('inf')),
                'good': (8, 15),
                'acceptable': (3, 8),
                'poor': (0, 3),
                'bad': (-float('inf'), 0)
            },
            'Growth 11M (%)': {
                'excellent': (15, float('inf')),
                'good': (8, 15),
                'acceptable': (3, 8),
                'poor': (0, 3),
                'bad': (-float('inf'), 0)
            },
            'Growth 12M (%)': {
                'excellent': (15, float('inf')),
                'good': (8, 15),
                'acceptable': (3, 8),
                'poor': (0, 3),
                'bad': (-float('inf'), 0)
            }
        }
    
    def _get_cell_color_for_value(self, metric_name, value):
        """
        Determine the appropriate cell color based on metric value and defined ranges.
        
        Args:
            metric_name (str): Name of the financial metric
            value: The numeric value to evaluate
            
        Returns:
            PatternFill: The appropriate color fill for the cell
        """
        if value is None or value == 'N/A' or value == '':
            return self.neutral_fill
            
        try:
            numeric_value = float(value)
        except (ValueError, TypeError):
            return self.neutral_fill
            
        # Get ranges for this metric
        ranges = self.metric_ranges.get(metric_name)
        if not ranges:
            # For unknown metrics, use simple threshold check
            return self._get_simple_threshold_color(metric_name, numeric_value)
            
        # Check which range the value falls into
        for performance, (min_val, max_val) in ranges.items():
            if min_val <= numeric_value < max_val:
                if performance == 'excellent':
                    return self.excellent_fill
                elif performance == 'good':
                    return self.good_fill
                elif performance == 'acceptable':
                    return self.acceptable_fill
                elif performance == 'poor':
                    return self.poor_fill
                elif performance == 'bad':
                    return self.bad_fill
                    
        return self.neutral_fill
    
    def _get_simple_threshold_color(self, metric_name, value):
        """
        Simple threshold-based coloring for metrics not in detailed ranges.
        
        Args:
            metric_name (str): Name of the metric
            value (float): Numeric value
            
        Returns:
            PatternFill: Appropriate color based on threshold comparison
        """
        # Map metric names to threshold keys
        threshold_map = {
            'PE Ratio': ('pe_ratio_max', '<'),
            'Debt/Equity': ('debt_to_equity_max', '<'), 
            'ROE': ('roe_min', '>'),
            'Current Ratio': ('current_ratio_min', '>'),
            'Price to Book': ('price_to_book_max', '<'),
            'Promoter Holding': ('promoter_holding_min', '>'),
            'Price to Cash Flow': ('price_to_cash_flow_max', '<'),
            'Quick Ratio': ('quick_ratio_min', '>'),
            'Interest Coverage Ratio': ('interest_coverage_min', '>'),
            'Free Cash Flow': ('free_cash_flow_min', '>'),
            'EPS Growth (%)': ('eps_growth_min', '>'),
            'Return on Assets (ROA)': ('roa_min', '>'),
            'Net Profit Margin': ('net_profit_margin_min', '>'),
            'Operating Margin': ('operating_margin_min', '>'),
            'Cash Conversion Ratio': ('cash_conversion_min', '>'),
            'EV/EBITDA': ('ev_ebitda_max', '<'),
            'Revenue Growth (%)': ('revenue_growth_min', '>'),
            'Pledged Shares (%)': ('pledged_shares_max', '<')
        }
        
        threshold_info = threshold_map.get(metric_name)
        if threshold_info:
            threshold_key, operator = threshold_info
            threshold_value = THRESHOLDS.get(threshold_key)
            
            if threshold_value is not None:
                if operator == '<' and value < threshold_value:
                    return self.good_fill
                elif operator == '>' and value > threshold_value:
                    return self.good_fill
                else:
                    return self.poor_fill
                    
        return self.neutral_fill

    def write_data(self, results: list[dict]):
        """
        Write the analysis results to an Excel file with enhanced conditional formatting.

        Args:
            results (list[dict]): List of analysis result dictionaries.
        """
        try:
            # Convert results into a DataFrame
            df = pd.DataFrame(results)
            logging.info(f"Initial DataFrame created with {len(df)} rows.")

            # Replace NaN values with "N/A" (handle dtype compatibility)
            df = df.fillna("N/A")

            # Identify metric columns (exclude non-metric ones)
            non_metric_cols = ['Symbol', 'Investment Recommendation']
            metric_columns = [col for col in df.columns if col not in non_metric_cols and not col.endswith(" Pass")]

            # Set recommendation to 'Insufficient data' if all metrics are 'N/A' or NaN
            for index, row in df.iterrows():
                all_na = all(row[col] in ['N/A', None, float('nan')] for col in metric_columns)
                if all_na:
                    df.at[index, 'Investment Recommendation'] = 'Insufficient data'
                    logging.info(f"Row {index} marked as 'Insufficient data' due to missing metrics.")

            # Ensure output directory exists
            os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
            # Save DataFrame to Excel
            df.to_excel(self.output_path, index=False)
            logging.info(f"Excel file saved at {self.output_path}")

            # Load workbook for advanced styling
            wb = load_workbook(self.output_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            # Remove *_Pass columns first (they're not needed for display)
            pass_columns = [col for col in headers if col.endswith("Pass") and col != "Investment Recommendation"]
            for col in reversed(pass_columns):  # Reverse to avoid index shifting
                col_index = headers.index(col) + 1
                ws.delete_cols(col_index)
                headers.pop(col_index - 1)
            
            # Refresh headers after column deletion
            headers = [cell.value for cell in ws[1]]
            logging.info(f"Headers after cleanup: {headers}")

            # Apply enhanced value-based conditional formatting
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                row_map = {headers[i]: cell for i, cell in enumerate(row)}

                # Apply precise color coding to each metric based on its value
                for metric_name in metric_columns:
                    if metric_name in row_map:
                        cell = row_map[metric_name]
                        if cell and cell.value is not None:
                            # Get appropriate color based on value performance
                            fill_color = self._get_cell_color_for_value(metric_name, cell.value)
                            cell.fill = fill_color
                            
                            # Add bold font for excellent performers
                            if fill_color == self.excellent_fill:
                                cell.font = self.bold_white_font
                            elif fill_color in [self.bad_fill, self.poor_fill]:
                                cell.font = self.bold_font

                # Enhanced Investment Recommendation formatting
                rec_cell = row_map.get("Investment Recommendation")
                if rec_cell and rec_cell.value:
                    rec_value = rec_cell.value
                    if rec_value == "Strong Buy":
                        rec_cell.fill = PatternFill(start_color='006100', end_color='006100', fill_type='solid')  # Dark Green
                        rec_cell.font = Font(bold=True, color='FFFFFF')
                    elif rec_value == "Buy":
                        rec_cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # Green
                        rec_cell.font = Font(bold=True, color='FFFFFF')
                    elif rec_value in ["Hold", "Weak Hold"]:
                        rec_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
                        rec_cell.font = Font(bold=True)
                    elif rec_value == "Avoid":
                        rec_cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')  # Dark Red
                        rec_cell.font = Font(bold=True, color='FFFFFF')
                    elif rec_value == "Insufficient data":
                        rec_cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # Gray
                        rec_cell.font = Font(bold=True, italic=True, color='FFFFFF')

                # Special formatting for Value Score
                value_score_cell = row_map.get("Value Score")
                if value_score_cell and value_score_cell.value is not None:
                    try:
                        score = float(value_score_cell.value)
                        fill_color = self._get_cell_color_for_value("Value Score", score)
                        value_score_cell.fill = fill_color
                        if fill_color in [self.excellent_fill, self.bad_fill]:
                            value_score_cell.font = Font(bold=True, color='FFFFFF')
                        else:
                            value_score_cell.font = Font(bold=True)
                    except (ValueError, TypeError):
                        pass

                # Comprehensive AI Analysis Formatting
                
                # Value Score (1-10) formatting
                value_score_10_cell = row_map.get("Value Score (1-10)")
                if value_score_10_cell and value_score_10_cell.value is not None:
                    try:
                        if str(value_score_10_cell.value).replace('.','').isdigit():
                            score = float(value_score_10_cell.value)
                            fill_color = self._get_cell_color_for_value("Value Score (1-10)", score)
                            value_score_10_cell.fill = fill_color
                            if fill_color in [self.excellent_fill, self.bad_fill]:
                                value_score_10_cell.font = Font(bold=True, color='FFFFFF')
                            else:
                                value_score_10_cell.font = Font(bold=True)
                    except (ValueError, TypeError):
                        pass

                # Financial Health formatting
                financial_health_cell = row_map.get("Financial Health")
                if financial_health_cell and financial_health_cell.value:
                    health_value = financial_health_cell.value
                    if health_value == "Excellent":
                        financial_health_cell.fill = self.excellent_fill
                        financial_health_cell.font = Font(bold=True, color='FFFFFF')
                    elif health_value == "Good":
                        financial_health_cell.fill = self.good_fill
                        financial_health_cell.font = Font(bold=True)
                    elif health_value == "Fair":
                        financial_health_cell.fill = self.acceptable_fill
                        financial_health_cell.font = Font(bold=True)
                    elif health_value == "Poor":
                        financial_health_cell.fill = self.poor_fill
                        financial_health_cell.font = Font(bold=True)

                # Business Quality formatting
                business_quality_cell = row_map.get("Business Quality")
                if business_quality_cell and business_quality_cell.value:
                    quality_value = business_quality_cell.value
                    if quality_value == "High":
                        business_quality_cell.fill = self.excellent_fill
                        business_quality_cell.font = Font(bold=True, color='FFFFFF')
                    elif quality_value == "Medium":
                        business_quality_cell.fill = self.acceptable_fill
                        business_quality_cell.font = Font(bold=True)
                    elif quality_value == "Low":
                        business_quality_cell.fill = self.poor_fill
                        business_quality_cell.font = Font(bold=True)

                # Risk Level formatting (inverse - lower risk is better)
                risk_level_cell = row_map.get("Risk Level")
                if risk_level_cell and risk_level_cell.value:
                    risk_value = risk_level_cell.value
                    if risk_value == "Low":
                        risk_level_cell.fill = self.excellent_fill
                        risk_level_cell.font = Font(bold=True, color='FFFFFF')
                    elif risk_value == "Medium":
                        risk_level_cell.fill = self.acceptable_fill
                        risk_level_cell.font = Font(bold=True)
                    elif risk_value == "High":
                        risk_level_cell.fill = self.poor_fill
                        risk_level_cell.font = Font(bold=True)
                    elif risk_value == "Very High":
                        risk_level_cell.fill = self.bad_fill
                        risk_level_cell.font = Font(bold=True, color='FFFFFF')

                # Valuation Assessment formatting
                valuation_cell = row_map.get("Valuation Assessment")
                if valuation_cell and valuation_cell.value:
                    valuation_value = valuation_cell.value
                    if valuation_value == "Undervalued":
                        valuation_cell.fill = self.excellent_fill
                        valuation_cell.font = Font(bold=True, color='FFFFFF')
                    elif valuation_value == "Fairly Valued":
                        valuation_cell.fill = self.good_fill
                        valuation_cell.font = Font(bold=True)
                    elif valuation_value == "Overvalued":
                        valuation_cell.fill = self.poor_fill
                        valuation_cell.font = Font(bold=True)

                # Margin of Safety formatting
                margin_safety_cell = row_map.get("Margin of Safety")
                if margin_safety_cell and margin_safety_cell.value:
                    margin_value = margin_safety_cell.value
                    if margin_value == "High":
                        margin_safety_cell.fill = self.excellent_fill
                        margin_safety_cell.font = Font(bold=True, color='FFFFFF')
                    elif margin_value == "Medium":
                        margin_safety_cell.fill = self.good_fill
                        margin_safety_cell.font = Font(bold=True)
                    elif margin_value == "Low":
                        margin_safety_cell.fill = self.acceptable_fill
                        margin_safety_cell.font = Font(bold=True)
                    elif margin_value == "None":
                        margin_safety_cell.fill = self.bad_fill
                        margin_safety_cell.font = Font(bold=True, color='FFFFFF')

                # AI Recommendation formatting (enhanced)
                ai_rec_cell = row_map.get("AI Recommendation")
                if ai_rec_cell and ai_rec_cell.value:
                    ai_rec_value = ai_rec_cell.value
                    if ai_rec_value == "Strong Buy":
                        ai_rec_cell.fill = PatternFill(start_color='004000', end_color='004000', fill_type='solid')  # Dark Green
                        ai_rec_cell.font = Font(bold=True, color='FFFFFF')
                    elif ai_rec_value == "Buy":
                        ai_rec_cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')  # Green
                        ai_rec_cell.font = Font(bold=True, color='FFFFFF')
                    elif ai_rec_value == "Hold":
                        ai_rec_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
                        ai_rec_cell.font = Font(bold=True)
                    elif ai_rec_value == "Sell":
                        ai_rec_cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')  # Orange
                        ai_rec_cell.font = Font(bold=True)
                    elif ai_rec_value == "Avoid":
                        ai_rec_cell.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')  # Dark Red
                        ai_rec_cell.font = Font(bold=True, color='FFFFFF')

                # AI Sentiment formatting
                ai_sentiment_cell = row_map.get("AI Sentiment")
                if ai_sentiment_cell and ai_sentiment_cell.value:
                    sentiment_value = ai_sentiment_cell.value
                    if sentiment_value == "Good":
                        ai_sentiment_cell.fill = self.good_fill
                        ai_sentiment_cell.font = Font(bold=True)
                    elif sentiment_value == "Neutral":
                        ai_sentiment_cell.fill = self.acceptable_fill
                        ai_sentiment_cell.font = Font(bold=True)
                    elif sentiment_value in ["Bad", "Worst"]:
                        ai_sentiment_cell.fill = self.poor_fill
                        ai_sentiment_cell.font = Font(bold=True)

            # Auto-adjust column widths for better readability
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)  # Cap at 25 characters
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(self.output_path)
            logging.info("Enhanced styling and formatting complete. File saved.")

        except Exception as e:
            logging.error(f"Error writing Excel file with enhanced formatting: {e}")
