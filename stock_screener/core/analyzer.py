#!/usr/bin/env python3
"""
Detailed Analysis Module for Stock Screener
============================================

This module provides enhanced analysis of screener output, focusing on stocks 
with "Buy" or "Strong Buy" recommendations for further detailed evaluation.

Features:
- Filters high-quality investment candidates
- Adds additional analysis metrics
- Creates enhanced Excel output with conditional formatting
- Provides investment readiness assessment
- Multi-provider AI analysis support

Usage:
    python detailed_analysis.py [input_file] [output_file] [options]
    
    Options:
    --no-ai                 Disable AI analysis (use rule-based only)
    --ai-provider=PROVIDER  Specify AI provider (auto, openai, gemini, claude, ollama)
    --openai               Use OpenAI GPT (shorthand for --ai-provider=openai)
    --gemini               Use Google Gemini (shorthand for --ai-provider=gemini)
    --claude               Use Anthropic Claude (shorthand for --ai-provider=claude)
    --ollama               Use local Ollama (shorthand for --ai-provider=ollama)
    
    Examples:
    python detailed_analysis.py                                    # Auto-detect AI provider
    python detailed_analysis.py --gemini                          # Force Google Gemini
    python detailed_analysis.py --ai-provider=openai              # Force OpenAI
    python detailed_analysis.py data/input.xlsx --claude          # Use Claude with custom input
    python detailed_analysis.py --no-ai                           # Disable AI analysis
    
    If no arguments provided, uses default paths:
    - Input: data/output/value_analysis.xlsx
    - Output: data/output/detailed_analysis.xlsx
"""

import pandas as pd
import sys
import os
from pathlib import Path
import logging
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import json
import time
import requests
from typing import Dict, Any, Optional

# AI/LLM dependencies - Support multiple providers
AI_PROVIDERS = {}

try:
    import openai
    AI_PROVIDERS['openai'] = True
except ImportError:
    AI_PROVIDERS['openai'] = False

try:
    import google.generativeai as genai
    AI_PROVIDERS['gemini'] = True
except ImportError:
    AI_PROVIDERS['gemini'] = False

try:
    import anthropic
    AI_PROVIDERS['claude'] = True
except ImportError:
    AI_PROVIDERS['claude'] = False

try:
    import requests  # For Ollama local API
    AI_PROVIDERS['ollama'] = True
except ImportError:
    AI_PROVIDERS['ollama'] = False

# Check if any AI provider is available
HAS_AI = any(AI_PROVIDERS.values())

try:
    from dotenv import load_dotenv
    load_dotenv()  # Load environment variables from .env file
except ImportError:
    pass  # dotenv not available, will use system environment variables

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Import price prediction service
try:
    from stock_screener.services.pricePrediction import PricePredictionService
    HAS_PRICE_PREDICTION = True
    logger.info("Price prediction service available")
except ImportError as e:
    HAS_PRICE_PREDICTION = False
    logger.warning(f"Price prediction service not available: {e}")


class DetailedAnalyzer:
    """
    Enhanced analyzer for stocks with Buy/Strong Buy recommendations.
    Provides additional insights, investment readiness assessment, and AI-powered analysis.
    """
    
    def __init__(self, input_file=None, output_file=None, enable_ai_analysis=True, preferred_ai_provider=None, config_file=None):
        """
        Initialize the detailed analyzer with input and output file paths.
        
        Args:
            input_file (str): Path to the screener output Excel file
            output_file (str): Path for the detailed analysis output
            enable_ai_analysis (bool): Enable LLM-powered analysis and news sentiment
            preferred_ai_provider (str): Preferred AI provider ('openai', 'gemini', 'claude', 'ollama', 'auto')
            config_file (str): Path to configuration file for AI settings
        """
        # Load configuration if provided
        if config_file and os.path.exists(config_file):
            import configparser
            config = configparser.ConfigParser()
            config.read(config_file)
            
            # Override settings with config values
            enable_ai_analysis = config.getboolean('DEFAULT', 'ai_analysis_enabled', fallback=enable_ai_analysis)
            preferred_ai_provider = config.get('DEFAULT', 'ai_provider', fallback=preferred_ai_provider)
            logger.info(f"AI settings loaded from config: enabled={enable_ai_analysis}, provider={preferred_ai_provider}")
        
        self.input_file = input_file or 'data/output/value_analysis.xlsx'
        self.output_file = output_file or 'data/output/detailed_analysis.xlsx'
        self.enable_ai_analysis = enable_ai_analysis and HAS_AI
        self.preferred_ai_provider = preferred_ai_provider or 'auto'
        
        # Initialize AI clients based on available providers
        self.ai_client = None
        self.ai_provider = None
        
        if self.enable_ai_analysis:
            self.ai_client, self.ai_provider = self._initialize_ai_client()
            if self.ai_client:
                logger.info(f"AI client initialized successfully using {self.ai_provider}")
            else:
                logger.warning("No AI provider available. AI analysis will be disabled.")
                self.enable_ai_analysis = False
        
        # Define investment readiness criteria
        self.readiness_weights = {
            'Value Score': 0.40,        # 40% weight for overall value
            'ROE': 0.15,               # 15% for profitability
            'Current Ratio': 0.10,     # 10% for liquidity
            'Debt/Equity': 0.10,       # 10% for financial health
            'PE Ratio': 0.15,          # 15% for valuation
            'Promoter Holding': 0.10   # 10% for management confidence
        }
        
        # Define ideal financial ratio ranges for AI analysis
        self.ideal_ratios = {
            'PE Ratio': {'excellent': (8, 15), 'good': (15, 20), 'acceptable': (20, 25)},
            'Price to Book': {'excellent': (0.5, 1.5), 'good': (1.5, 2.5), 'acceptable': (2.5, 3.5)},
            'ROE': {'excellent': (20, 50), 'good': (15, 20), 'acceptable': (10, 15)},
            'Current Ratio': {'excellent': (2.0, 3.0), 'good': (1.5, 2.0), 'acceptable': (1.0, 1.5)},
            'Debt/Equity': {'excellent': (0, 0.3), 'good': (0.3, 0.6), 'acceptable': (0.6, 1.0)},
            'Net Profit Margin': {'excellent': (15, 50), 'good': (10, 15), 'acceptable': (5, 10)},
            'EV/EBITDA': {'excellent': (5, 10), 'good': (10, 15), 'acceptable': (15, 20)},
            'Promoter Holding': {'excellent': (50, 75), 'good': (40, 50), 'acceptable': (25, 40)}
        }
        
    def _initialize_ai_client(self):
        """
        Initialize AI client from available providers based on preference.
        
        Returns:
            tuple: (client_object, provider_name) or (None, None) if no provider available
        """
        logger.info(f"AI Provider Preference: {self.preferred_ai_provider}")
        
        # If specific provider is requested, try only that one
        if self.preferred_ai_provider != 'auto':
            if self.preferred_ai_provider == 'openai':
                return self._try_openai()
            elif self.preferred_ai_provider == 'gemini':
                return self._try_gemini()
            elif self.preferred_ai_provider == 'claude':
                return self._try_claude()
            elif self.preferred_ai_provider == 'ollama':
                return self._try_ollama()
            else:
                logger.warning(f"Unknown AI provider: {self.preferred_ai_provider}. Falling back to auto-detection.")
        
        # Auto-detection mode: Try providers in default priority order
        logger.info("Auto-detecting available AI providers...")
        
        # Priority order: OpenAI (reliable) > Gemini (free but limited) > Claude > Ollama (local)
        
        # 1. Try OpenAI (higher rate limits)
        result = self._try_openai()
        if result[0]:
            return result
        
        # 2. Try Google Gemini (FREE tier available but rate limited)
        result = self._try_gemini()
        if result[0]:
            return result
        
        # 3. Try Anthropic Claude
        result = self._try_claude()
        if result[0]:
            return result
        
        # 4. Try Ollama (local - no API key needed)
        result = self._try_ollama()
        if result[0]:
            return result
        
        # 5. Fallback to rule-based analysis
        logger.info("No AI provider available, using enhanced rule-based analysis")
        return 'rule_based', 'rule_based'
    
    def _try_openai(self):
        """Try to initialize OpenAI client."""
        if AI_PROVIDERS.get('openai') and os.getenv('OPENAI_API_KEY'):
            try:
                logger.info("Attempting to initialize OpenAI client...")
                return openai.OpenAI(api_key=os.getenv('OPENAI_API_KEY')), 'openai'
            except Exception as e:
                logger.warning(f"Failed to initialize OpenAI: {e}")
        else:
            logger.debug("OpenAI not available (missing package or API key)")
        return None, None
    
    def _try_gemini(self):
        """Try to initialize Google Gemini client."""
        if AI_PROVIDERS.get('gemini') and (os.getenv('GOOGLE_API_KEY') or os.getenv('GEMINI_API_KEY')):
            try:
                logger.info("Attempting to initialize Google Gemini client...")
                api_key = os.getenv('GOOGLE_API_KEY') or os.getenv('GEMINI_API_KEY')
                genai.configure(api_key=api_key)
                return genai.GenerativeModel('gemini-1.5-flash'), 'gemini'
            except Exception as e:
                logger.warning(f"Failed to initialize Gemini: {e}")
        else:
            logger.debug("Gemini not available (missing package or API key)")
        return None, None
    
    def _try_claude(self):
        """Try to initialize Anthropic Claude client."""
        if AI_PROVIDERS.get('claude') and os.getenv('ANTHROPIC_API_KEY'):
            try:
                logger.info("Attempting to initialize Anthropic Claude client...")
                return anthropic.Anthropic(api_key=os.getenv('ANTHROPIC_API_KEY')), 'claude'
            except Exception as e:
                logger.warning(f"Failed to initialize Claude: {e}")
        else:
            logger.debug("Claude not available (missing package or API key)")
        return None, None
    
    def _try_ollama(self):
        """Try to initialize Ollama local client."""
        if AI_PROVIDERS.get('ollama'):
            try:
                logger.info("Attempting to connect to local Ollama server...")
                # Test if Ollama is running locally
                response = requests.get('http://localhost:11434/api/tags', timeout=2)
                if response.status_code == 200:
                    return 'ollama_client', 'ollama'
                else:
                    logger.debug(f"Ollama server returned status {response.status_code}")
            except Exception as e:
                logger.debug(f"Ollama not available locally: {e}")
        else:
            logger.debug("Ollama not available (missing requests package)")
        return None, None
        
    def load_screener_data(self):
        """
        Load and validate the screener output data.
        
        Returns:
            pd.DataFrame: Loaded screener data
            
        Raises:
            FileNotFoundError: If input file doesn't exist
            ValueError: If required columns are missing
        """
        if not os.path.exists(self.input_file):
            raise FileNotFoundError(f"Input file not found: {self.input_file}")
            
        logger.info(f"Loading screener data from: {self.input_file}")
        df = pd.read_excel(self.input_file)
        
        # Validate required columns
        required_columns = ['Symbol', 'Investment Recommendation', 'Value Score']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")
            
        logger.info(f"Loaded {len(df)} stocks from screener output")
        return df
    
    def filter_investment_candidates(self, df):
        """
        Filter stocks with 'Buy' or 'Strong Buy' recommendations.
        
        Args:
            df (pd.DataFrame): Input screener data
            
        Returns:
            pd.DataFrame: Filtered data with buy recommendations
        """
        buy_recommendations = ['Buy', 'Strong Buy']
        filtered_df = df[df['Investment Recommendation'].isin(buy_recommendations)].copy()
        
        logger.info(f"Found {len(filtered_df)} stocks with Buy/Strong Buy recommendations:")
        for rec in buy_recommendations:
            count = len(filtered_df[filtered_df['Investment Recommendation'] == rec])
            if count > 0:
                logger.info(f"  - {rec}: {count} stocks")
                
        return filtered_df
    
    def calculate_investment_readiness(self, df):
        """
        Calculate an Investment Readiness Score based on key metrics.
        
        Args:
            df (pd.DataFrame): Filtered investment candidates
            
        Returns:
            pd.DataFrame: Data with added Investment Readiness Score
        """
        logger.info("Calculating Investment Readiness Scores...")
        
        def score_metric(value, metric_name, is_higher_better=True):
            """Score individual metrics on 0-10 scale."""
            if pd.isna(value) or value == 'N/A':
                return 5.0  # Neutral score for missing data
                
            try:
                value = float(value)
            except (ValueError, TypeError):
                return 5.0
                
            if metric_name == 'Value Score':
                # Value Score is already 0-100, convert to 0-10
                return min(10.0, max(0.0, value / 10.0))
                
            elif metric_name == 'ROE':
                if value >= 25: return 10.0
                elif value >= 20: return 9.0
                elif value >= 15: return 8.0
                elif value >= 10: return 6.0
                elif value >= 5: return 4.0
                return 2.0
                
            elif metric_name == 'Current Ratio':
                if value >= 2.5: return 10.0
                elif value >= 2.0: return 9.0
                elif value >= 1.5: return 8.0
                elif value >= 1.0: return 6.0
                return 3.0
                
            elif metric_name == 'Debt/Equity':
                # Lower is better for debt
                if value <= 0.3: return 10.0
                elif value <= 0.5: return 9.0
                elif value <= 0.8: return 8.0
                elif value <= 1.0: return 6.0
                elif value <= 1.5: return 4.0
                return 2.0
                
            elif metric_name == 'PE Ratio':
                # Lower is better for PE
                if value <= 0: return 0.0
                elif value <= 10: return 10.0
                elif value <= 15: return 9.0
                elif value <= 20: return 8.0
                elif value <= 25: return 6.0
                elif value <= 30: return 4.0
                return 2.0
                
            elif metric_name == 'Promoter Holding':
                if value >= 70: return 10.0
                elif value >= 60: return 9.0
                elif value >= 50: return 8.0
                elif value >= 40: return 6.0
                elif value >= 25: return 4.0
                return 2.0
                
            return 5.0  # Default neutral score
        
        # Calculate weighted readiness score
        readiness_scores = []
        for _, row in df.iterrows():
            total_score = 0.0
            total_weight = 0.0
            
            for metric, weight in self.readiness_weights.items():
                if metric in df.columns:
                    score = score_metric(row[metric], metric)
                    total_score += score * weight
                    total_weight += weight
                    
            # Normalize to 0-100 scale
            final_score = (total_score / total_weight) * 10 if total_weight > 0 else 50.0
            readiness_scores.append(round(final_score, 1))
        
        df['Investment Readiness Score'] = readiness_scores
        
        # Add readiness category
        def categorize_readiness(score):
            if score >= 80: return 'Excellent'
            elif score >= 70: return 'Very Good'
            elif score >= 60: return 'Good'
            elif score >= 50: return 'Fair'
            else: return 'Needs Review'
            
        df['Readiness Category'] = df['Investment Readiness Score'].apply(categorize_readiness)
        
        logger.info(f"Added Investment Readiness Scores (avg: {df['Investment Readiness Score'].mean():.1f})")
        return df
    
    def search_stock_news(self, symbol: str, days_back: int = 30) -> str:
        """
        Search for recent news about a stock symbol using multiple real news sources.
        
        Args:
            symbol (str): Stock symbol (e.g., 'TCS.NS')
            days_back (int): Number of days to search back for news
            
        Returns:
            str: Recent news summary with real data or empty string if no news found
        """
        try:
            # Clean symbol for search (remove .NS suffix)
            clean_symbol = symbol.replace('.NS', '').replace('.BO', '')
            company_name = self._get_company_name(clean_symbol)
            
            logger.info(f"Searching real news for {symbol} ({company_name})...")
            
            news_sources = []
            
            # Method 1: Try to get news from financial websites
            news_sources.extend(self._scrape_financial_news(clean_symbol, company_name))
            
            # Method 2: Try News API if available
            if os.getenv('NEWS_API_KEY'):
                news_sources.extend(self._fetch_news_api(clean_symbol, company_name, days_back))
            
            # Method 3: Try Google News RSS
            news_sources.extend(self._fetch_google_news_rss(clean_symbol, company_name))
            
            # Compile news summary
            if news_sources:
                # Sort by relevance and recency
                news_sources = news_sources[:5]  # Top 5 most relevant
                
                news_summary = f"Recent news for {company_name} ({clean_symbol}):\n"
                for i, news in enumerate(news_sources, 1):
                    news_summary += f"{i}. {news['title']} ({news['source']}) - {news['summary']}\n"
                
                logger.info(f"Found {len(news_sources)} recent news items for {symbol}")
                return news_summary
            else:
                logger.warning(f"No recent news found for {symbol}")
                return f"No recent significant news found for {company_name} in the last {days_back} days"
            
        except Exception as e:
            logger.warning(f"News search failed for {symbol}: {e}")
            return f"News analysis unavailable for {symbol}"
    
    def _get_company_name(self, symbol: str) -> str:
        """Get full company name from symbol."""
        # Common Indian stock symbols to company names mapping
        symbol_to_name = {
            'ITC': 'ITC Limited',
            'TCS': 'Tata Consultancy Services',
            'INFY': 'Infosys Limited',
            'HDFC': 'HDFC Bank',
            'HDFCBANK': 'HDFC Bank',
            'RELIANCE': 'Reliance Industries',
            'SBIN': 'State Bank of India',
            'WIPRO': 'Wipro Limited',
            'ONGC': 'Oil and Natural Gas Corporation',
            'HINDUNILVR': 'Hindustan Unilever',
            'ICICIBANK': 'ICICI Bank',
            'AXISBANK': 'Axis Bank',
            'LT': 'Larsen & Toubro',
            'BHARTIARTL': 'Bharti Airtel',
            'KOTAKBANK': 'Kotak Mahindra Bank',
            'ASIANPAINT': 'Asian Paints',
            'M&M': 'Mahindra & Mahindra',
            'TATAMOTORS': 'Tata Motors',
            'MARUTI': 'Maruti Suzuki',
            'BAJFINANCE': 'Bajaj Finance'
        }
        return symbol_to_name.get(symbol.upper(), symbol + ' Limited')
    
    def _scrape_financial_news(self, symbol: str, company_name: str) -> list:
        """Scrape recent news from financial websites."""
        news_items = []
        
        try:
            # Try to get news from multiple financial sources
            import requests
            from bs4 import BeautifulSoup
            import re
            from urllib.parse import quote
            
            # Source 1: Economic Times
            try:
                search_url = f"https://economictimes.indiatimes.com/topic/{quote(company_name)}"
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                }
                
                response = requests.get(search_url, headers=headers, timeout=10)
                if response.status_code == 200:
                    soup = BeautifulSoup(response.content, 'html.parser')
                    
                    # Find news headlines
                    headlines = soup.find_all(['h2', 'h3'], class_=re.compile(r'title|headline'), limit=3)
                    
                    for headline in headlines:
                        title = headline.get_text().strip()
                        if len(title) > 20 and any(word in title.lower() for word in [symbol.lower(), company_name.split()[0].lower()]):
                            news_items.append({
                                'title': title[:100],
                                'source': 'Economic Times',
                                'summary': f'Recent development regarding {company_name}',
                                'relevance': 0.9
                            })
                            
            except Exception as e:
                logger.debug(f"Economic Times scraping failed: {e}")
            
            # Source 2: MoneyControl  
            try:
                search_url = f"https://www.moneycontrol.com/news/tags/{symbol.lower()}.html"
                response = requests.get(search_url, headers=headers, timeout=10)
                
                if response.status_code == 200:
                    soup = BeautifulSoup(response.content, 'html.parser')
                    headlines = soup.find_all('h2', limit=2)
                    
                    for headline in headlines:
                        title = headline.get_text().strip()
                        if len(title) > 20:
                            news_items.append({
                                'title': title[:100], 
                                'source': 'MoneyControl',
                                'summary': f'Financial news about {company_name}',
                                'relevance': 0.8
                            })
                            
            except Exception as e:
                logger.debug(f"MoneyControl scraping failed: {e}")
                
        except ImportError:
            logger.debug("BeautifulSoup not available for web scraping")
        except Exception as e:
            logger.debug(f"Web scraping failed: {e}")
            
        return news_items[:3]  # Return top 3 items
    
    def _fetch_news_api(self, symbol: str, company_name: str, days_back: int) -> list:
        """Fetch news using News API if available."""
        news_items = []
        
        try:
            import requests
            from datetime import datetime, timedelta
            
            api_key = os.getenv('NEWS_API_KEY')
            if not api_key:
                return news_items
                
            # Calculate date range
            to_date = datetime.now()
            from_date = to_date - timedelta(days=days_back)
            
            # Search for company news
            url = "https://newsapi.org/v2/everything"
            params = {
                'q': f'"{company_name}" OR "{symbol}" AND (earnings OR results OR stock OR shares OR profit OR revenue)',
                'language': 'en',
                'sortBy': 'publishedAt',
                'pageSize': 3,
                'from': from_date.strftime('%Y-%m-%d'),
                'to': to_date.strftime('%Y-%m-%d'),
                'apiKey': api_key
            }
            
            response = requests.get(url, params=params, timeout=10)
            
            if response.status_code == 200:
                data = response.json()
                
                for article in data.get('articles', [])[:3]:
                    news_items.append({
                        'title': article.get('title', '')[:100],
                        'source': article.get('source', {}).get('name', 'News API'),
                        'summary': article.get('description', '')[:200],
                        'relevance': 0.95  # High relevance from dedicated API
                    })
                    
            logger.debug(f"News API returned {len(news_items)} articles")
                    
        except Exception as e:
            logger.debug(f"News API fetch failed: {e}")
            
        return news_items
    
    def _fetch_google_news_rss(self, symbol: str, company_name: str) -> list:
        """Fetch news from Google News RSS."""
        news_items = []
        
        try:
            import requests
            import xml.etree.ElementTree as ET
            from urllib.parse import quote
            
            # Google News RSS search
            query = f'{company_name} stock earnings financial results'
            rss_url = f"https://news.google.com/rss/search?q={quote(query)}&hl=en-IN&gl=IN&ceid=IN:en"
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            
            response = requests.get(rss_url, headers=headers, timeout=10)
            
            if response.status_code == 200:
                root = ET.fromstring(response.content)
                
                # Parse RSS items
                items = root.findall('.//item')[:2]  # Top 2 items
                
                for item in items:
                    title = item.find('title').text if item.find('title') is not None else ''
                    description = item.find('description').text if item.find('description') is not None else ''
                    
                    if len(title) > 10:
                        news_items.append({
                            'title': title[:100],
                            'source': 'Google News',
                            'summary': description[:150] if description else f'News about {company_name}',
                            'relevance': 0.7
                        })
                        
            logger.debug(f"Google News RSS returned {len(news_items)} articles")
                        
        except Exception as e:
            logger.debug(f"Google News RSS fetch failed: {e}")
            
        return news_items
    
    def analyze_stock_with_ai(self, row_data: Dict[str, Any]) -> Dict[str, str]:
        """
        Use AI (multiple providers) to analyze stock data against ideal ratios and provide recommendations.
        
        Args:
            row_data (dict): Stock data row containing all financial metrics
            
        Returns:
            dict: AI analysis results with sentiment, recommendation, and price predictions
        """
        if not self.enable_ai_analysis or not self.ai_client:
            return {
                'ai_sentiment': 'Not Available',
                'ai_recommendation': 'Standard Analysis',
                'ai_reasoning': 'AI analysis not available'
            }
        
        try:
            symbol = row_data.get('Symbol', 'Unknown')
            
            # Get recent news (simplified for now)
            news_summary = self.search_stock_news(symbol)
            
            # Get price predictions if available
            price_prediction = self._get_price_prediction(symbol)
            
            # Prepare financial data for AI analysis
            financial_metrics = {}
            for metric in ['PE Ratio', 'Price to Book', 'ROE', 'Current Ratio', 
                          'Debt/Equity', 'Net Profit Margin', 'EV/EBITDA', 'Promoter Holding']:
                if metric in row_data:
                    financial_metrics[metric] = row_data[metric]
            
            # Enhanced AI prompt with price prediction context
            prompt = self._create_enhanced_ai_analysis_prompt(symbol, financial_metrics, news_summary, price_prediction)
            
            # Get AI analysis result
            ai_result = {}
            if self.ai_provider == 'rule_based':
                ai_result = self._rule_based_analysis(financial_metrics)
            elif self.ai_provider == 'gemini':
                ai_result = self._call_gemini_api(prompt)
            elif self.ai_provider == 'openai':
                ai_result = self._call_openai_api(prompt)
            elif self.ai_provider == 'claude':
                ai_result = self._call_claude_api(prompt)
            elif self.ai_provider == 'ollama':
                ai_result = self._call_ollama_api(prompt)
            else:
                ai_result = self._rule_based_analysis(financial_metrics)
            
            # Enhance result with price prediction data
            if price_prediction and not price_prediction.get('error'):
                ai_result['predicted_price'] = price_prediction.get('predicted_price', 'N/A')
                ai_result['price_change_percent'] = price_prediction.get('price_change_percent', 'N/A')
                ai_result['prediction_confidence'] = price_prediction.get('confidence', 'N/A')
                ai_result['prediction_method'] = price_prediction.get('method', 'N/A')
                
                # Include multi-period predictions if available
                if 'multi_period_predictions' in price_prediction:
                    ai_result['multi_period_predictions'] = price_prediction['multi_period_predictions']
                
                # Override target price if we have a better prediction
                if 'target_price' not in ai_result or ai_result['target_price'] == 'N/A':
                    ai_result['target_price'] = str(price_prediction.get('predicted_price', 'N/A'))
            
            return ai_result
            
        except Exception as e:
            logger.error(f"AI analysis failed for {symbol}: {e}")
            return {
                'ai_sentiment': 'Error',
                'ai_recommendation': 'Analysis Failed',
                'ai_reasoning': f'Technical error: {str(e)[:100]}'
            }
    
    def _get_price_prediction(self, symbol: str) -> Dict:
        """Get comprehensive price prediction including multi-period forecasts for the stock"""
        if not HAS_PRICE_PREDICTION:
            return {"error": "Price prediction service not available"}
        
        try:
            predictor = PricePredictionService(symbol, prediction_days=30)
            
            # Get both single-period comprehensive predictions and multi-period predictions
            comprehensive_results = predictor.get_comprehensive_predictions()
            multi_period_results = predictor.get_simplified_multi_period_predictions()
            
            # Extract the ensemble result for the main output (30-day prediction)
            main_prediction = {}
            if "ensemble" in comprehensive_results and "predicted_price" in comprehensive_results["ensemble"]:
                ensemble = comprehensive_results["ensemble"]
                current_price = comprehensive_results.get("current_price", 0)
                predicted_price = ensemble["predicted_price"]
                
                # Calculate price change percentage
                price_change_percent = 0
                if current_price > 0:
                    price_change_percent = round((predicted_price - current_price) / current_price * 100, 2)
                
                main_prediction = {
                    "symbol": symbol,
                    "current_price": current_price,
                    "predicted_price": predicted_price,
                    "price_change_percent": price_change_percent,
                    "confidence": ensemble["confidence"],
                    "method": "Comprehensive Prediction (7 Methods + Ensemble)",
                    "methods_used": list(comprehensive_results["methods"].keys()),
                    "risk_level": comprehensive_results.get("risk_assessment", {}).get("risk_level", "Medium"),
                    "recommendation": ensemble.get("recommendation", "Hold")
                }
                
                # Add multi-period predictions if available
                if "predictions" in multi_period_results:
                    main_prediction["multi_period_predictions"] = multi_period_results["predictions"]
                    main_prediction["multi_period_summary"] = multi_period_results.get("summary", {})
                
                logger.info(f"Comprehensive price prediction for {symbol}: {main_prediction.get('predicted_price', 'N/A')} (Confidence: {main_prediction.get('confidence', 'N/A')})")
                return main_prediction
            else:
                # Fallback to quick prediction if comprehensive fails
                prediction = predictor.get_quick_prediction()
                logger.warning(f"Fell back to quick prediction for {symbol}")
                return prediction
        except Exception as e:
            logger.error(f"Price prediction failed for {symbol}: {e}")
            return {"error": f"Price prediction failed: {str(e)}"}
    
    def _create_enhanced_ai_analysis_prompt(self, symbol: str, metrics: Dict[str, Any], news: str, price_prediction: Dict) -> str:
        """Create an enhanced AI analysis prompt with price prediction context"""
        
        # Start with the basic prompt
        basic_prompt = self._create_ai_analysis_prompt(symbol, metrics, news)
        
        # Add price prediction context if available
        if price_prediction and not price_prediction.get('error'):
            prediction_context = f"""

PRICE PREDICTION ANALYSIS:
Current Price: ₹{price_prediction.get('current_price', 'N/A')}
Predicted Price (30 days): ₹{price_prediction.get('predicted_price', 'N/A')}
Expected Change: {price_prediction.get('price_change_percent', 'N/A')}%
Prediction Confidence: {price_prediction.get('confidence', 0)*100:.1f}%
Prediction Method: {price_prediction.get('method', 'N/A')}

ENHANCED ANALYSIS REQUIRED:
Based on the price prediction analysis above, provide enhanced insights considering:
1. How the predicted price movement aligns with fundamental analysis
2. Whether current valuation supports the predicted direction
3. Risk factors that could impact the price prediction accuracy
4. Timeline considerations for achieving the target price

Update your TARGET_PRICE recommendation considering this technical/fundamental price prediction."""
            
            return basic_prompt + prediction_context
        
        return basic_prompt
    
    def _create_ai_analysis_prompt(self, symbol: str, metrics: Dict[str, Any], news: str) -> str:
        """Create a comprehensive AI analysis prompt for value investing."""
        
        # Create detailed ratio analysis section
        ratio_analysis = []
        for metric, value in metrics.items():
            if metric in self.ideal_ratios and value is not None and value != 'N/A':
                try:
                    numeric_value = float(value)
                    ideal_ranges = self.ideal_ratios[metric]
                    
                    # Determine if metric is in ideal range
                    category = 'poor'
                    for cat, (min_val, max_val) in ideal_ranges.items():
                        if min_val <= numeric_value <= max_val:
                            category = cat
                            break
                    
                    ratio_analysis.append(f"- {metric}: {numeric_value} ({category} range)")
                except (ValueError, TypeError):
                    ratio_analysis.append(f"- {metric}: {value} (data issue)")
            else:
                ratio_analysis.append(f"- {metric}: {value}")
        
        prompt = f"""
Conduct a comprehensive value investing analysis for this Indian stock based on REAL DATA:

COMPANY: {symbol}

FINANCIAL METRICS:
{chr(10).join(ratio_analysis)}

RECENT NEWS & MARKET DEVELOPMENTS:
{news if news else "Limited recent news available - analysis based primarily on financial metrics"}

CRITICAL INSTRUCTION: Base your analysis on REAL FACTS from the financial metrics and news provided above. Do NOT make assumptions about data not provided. If recent news indicates specific developments (earnings, partnerships, regulatory changes, management changes, etc.), factor these into your analysis.

VALUE INVESTING BENCHMARKS (Graham & Buffett Principles):
- PE Ratio: Excellent (8-15), Good (15-20), Acceptable (20-25), Poor (>25)
- Price to Book: Excellent (0.5-1.5), Good (1.5-2.5), Acceptable (2.5-3.5), Poor (>3.5)
- ROE: Excellent (20-50%), Good (15-20%), Acceptable (10-15%), Poor (<10%)
- Current Ratio: Excellent (2.0-3.0), Good (1.5-2.0), Acceptable (1.0-1.5), Poor (<1.0)
- Debt/Equity: Excellent (0-0.3), Good (0.3-0.6), Acceptable (0.6-1.0), Poor (>1.0)

Provide analysis covering:

**1. CURRENT SITUATION ASSESSMENT**
- Recent news impact on business fundamentals
- Financial metric analysis vs benchmarks
- Market sentiment from recent developments

**2. FINANCIAL HEALTH REALITY CHECK** 
- Actual liquidity position based on available data
- Real debt levels and their sustainability
- Profitability trends evident from metrics
- Warning signs or positive indicators

**3. BUSINESS QUALITY ANALYSIS**
- Competitive position based on news/developments
- Management effectiveness (if mentioned in news)
- Industry challenges/opportunities from recent news
- Revenue quality assessment

**4. RISK ASSESSMENT BASED ON FACTS**
- Immediate risks from recent developments
- Financial risks evident from metrics
- Market/industry risks from news context
- Regulatory or competitive threats mentioned

**5. VALUATION REALITY**
- Current valuation vs intrinsic value using available metrics
- Margin of safety calculation based on real data
- Recent market developments impact on fair value

**6. INVESTMENT DECISION**
- Clear recommendation based on facts provided
- Specific reasons backed by data/news
- Timeline and conditions for the recommendation

Format your response as:
VALUE_SCORE: [1-10]
FINANCIAL_HEALTH: [Excellent/Good/Fair/Poor]
BUSINESS_QUALITY: [High/Medium/Low]
RISK_LEVEL: [Low/Medium/High/Very High]
VALUATION: [Undervalued/Fairly Valued/Overvalued]
MARGIN_OF_SAFETY: [High/Medium/Low/None]
RECOMMENDATION: [Strong Buy/Buy/Hold/Sell/Avoid]
INVESTMENT_THESIS: [Factual 4-5 sentence analysis based on provided data and news]
KEY_RISKS: [Specific risks from actual data/news, not generic risks]
CATALYSTS: [Real upcoming events or developments mentioned in news]
TARGET_PRICE: [Fair value estimate based on available metrics]

IMPORTANT: 
1. Use EXACTLY the format above. Do NOT use markdown.
2. Base analysis ONLY on provided financial data and news.
3. If news mentions specific developments, incorporate them into reasoning.
4. Avoid generic statements - use specific facts from the data provided.
"""
        return prompt
    
    def _parse_ai_response(self, response: str) -> Dict[str, str]:
        """Parse the comprehensive AI response into structured data."""
        try:
            lines = response.strip().split('\n')
            result = {
                'ai_sentiment': 'Neutral',
                'ai_recommendation': 'Hold', 
                'ai_reasoning': 'Analysis completed',
                'value_score': 'N/A',
                'financial_health': 'N/A',
                'business_quality': 'N/A',
                'risk_level': 'N/A',
                'valuation': 'N/A',
                'margin_of_safety': 'N/A',
                'investment_thesis': 'N/A',
                'key_risks': 'N/A',
                'catalysts': 'N/A',
                'target_price': 'N/A'
            }
            
            for line in lines:
                line = line.strip()
                
                # Parse comprehensive analysis fields
                if line.startswith('VALUE_SCORE:'):
                    score = line.replace('VALUE_SCORE:', '').strip()
                    result['value_score'] = score
                    # Convert to sentiment for backward compatibility
                    try:
                        score_num = float(score)
                        if score_num >= 7:
                            result['ai_sentiment'] = 'Good'
                        elif score_num >= 4:
                            result['ai_sentiment'] = 'Neutral'
                        else:
                            result['ai_sentiment'] = 'Bad'
                    except:
                        result['ai_sentiment'] = 'Neutral'
                        
                elif line.startswith('FINANCIAL_HEALTH:'):
                    result['financial_health'] = line.replace('FINANCIAL_HEALTH:', '').strip()
                    
                elif line.startswith('BUSINESS_QUALITY:'):
                    result['business_quality'] = line.replace('BUSINESS_QUALITY:', '').strip()
                    
                elif line.startswith('RISK_LEVEL:'):
                    result['risk_level'] = line.replace('RISK_LEVEL:', '').strip()
                    
                elif line.startswith('VALUATION:'):
                    result['valuation'] = line.replace('VALUATION:', '').strip()
                    
                elif line.startswith('MARGIN_OF_SAFETY:'):
                    result['margin_of_safety'] = line.replace('MARGIN_OF_SAFETY:', '').strip()
                    
                elif line.startswith('RECOMMENDATION:'):
                    recommendation = line.replace('RECOMMENDATION:', '').strip()
                    if recommendation in ['Strong Buy', 'Buy', 'Hold', 'Sell', 'Avoid']:
                        result['ai_recommendation'] = recommendation
                        
                elif line.startswith('INVESTMENT_THESIS:'):
                    thesis = line.replace('INVESTMENT_THESIS:', '').strip()
                    if thesis:
                        result['investment_thesis'] = thesis
                        result['ai_reasoning'] = thesis  # For backward compatibility
                        
                elif line.startswith('KEY_RISKS:'):
                    result['key_risks'] = line.replace('KEY_RISKS:', '').strip()
                    
                elif line.startswith('CATALYSTS:'):
                    result['catalysts'] = line.replace('CATALYSTS:', '').strip()
                    
                elif line.startswith('TARGET_PRICE:'):
                    result['target_price'] = line.replace('TARGET_PRICE:', '').strip()
                
                # Legacy format support
                elif line.startswith('SENTIMENT:'):
                    sentiment = line.replace('SENTIMENT:', '').strip()
                    if sentiment in ['Good', 'Bad', 'Worst']:
                        result['ai_sentiment'] = sentiment
                        
                elif line.startswith('REASONING:'):
                    reasoning = line.replace('REASONING:', '').strip()
                    if reasoning and result['ai_reasoning'] == 'Analysis completed':
                        result['ai_reasoning'] = reasoning
            
            return result
            
        except Exception as e:
            logger.error(f"Failed to parse AI response: {e}")
            return {
                'ai_sentiment': 'Error',
                'ai_recommendation': 'Hold',
                'ai_reasoning': 'Response parsing failed',
                'value_score': 'N/A',
                'financial_health': 'N/A',
                'business_quality': 'N/A',
                'risk_level': 'N/A',
                'valuation': 'N/A',
                'margin_of_safety': 'N/A',
                'investment_thesis': 'N/A',
                'key_risks': 'N/A',
                'catalysts': 'N/A',
                'target_price': 'N/A'
            }
    
    def _rule_based_analysis(self, metrics: Dict[str, Any]) -> Dict[str, str]:
        """
        Enhanced rule-based analysis when no AI provider is available.
        """
        excellent_count = 0
        good_count = 0
        poor_count = 0
        
        concerns = []
        strengths = []
        
        # Analyze each metric
        for metric, value in metrics.items():
            if metric in self.ideal_ratios and value is not None and value != 'N/A':
                try:
                    numeric_value = float(value)
                    ranges = self.ideal_ratios[metric]
                    
                    if 'excellent' in ranges:
                        min_ex, max_ex = ranges['excellent']
                        if min_ex <= numeric_value <= max_ex:
                            excellent_count += 1
                            strengths.append(f"Excellent {metric}")
                            continue
                    
                    if 'good' in ranges:
                        min_good, max_good = ranges['good']  
                        if min_good <= numeric_value <= max_good:
                            good_count += 1
                            continue
                    
                    poor_count += 1
                    concerns.append(f"Concerning {metric}")
                    
                except (ValueError, TypeError):
                    continue
        
        # Determine sentiment
        if excellent_count >= 4:
            sentiment = 'Good'
        elif excellent_count + good_count >= 5:
            sentiment = 'Good' 
        elif poor_count >= 4:
            sentiment = 'Worst'
        else:
            sentiment = 'Bad'
        
        # Determine recommendation
        if excellent_count >= 5:
            recommendation = 'Strong Buy'
        elif excellent_count >= 3:
            recommendation = 'Buy'
        elif poor_count >= 5:
            recommendation = 'Avoid'
        else:
            recommendation = 'Hold'
        
        # Create reasoning
        if strengths:
            strength_text = f"Strong metrics: {', '.join(strengths[:2])}"
        else:
            strength_text = "Mixed financial indicators"
            
        if concerns:
            concern_text = f"Areas of concern: {', '.join(concerns[:2])}"
            reasoning = f"{strength_text}. {concern_text}."
        else:
            reasoning = f"{strength_text}. Generally solid fundamentals."
        
        return {
            'ai_sentiment': sentiment,
            'ai_recommendation': recommendation,
            'ai_reasoning': reasoning[:150]  # Limit length
        }
    
    def _call_gemini_api(self, prompt: str) -> Dict[str, str]:
        """Call Google Gemini API with enhanced configuration for comprehensive analysis."""
        try:
            # Configure generation parameters for comprehensive analysis
            generation_config = {
                "temperature": 0.2,
                "top_p": 0.95,
                "top_k": 40,
                "max_output_tokens": 1500,  # Increased for comprehensive analysis
            }
            
            response = self.ai_client.generate_content(
                prompt,
                generation_config=generation_config
            )
            return self._parse_ai_response(response.text)
        except Exception as e:
            logger.error(f"Gemini API call failed: {e}")
            return self._rule_based_analysis({})
    
    def _call_openai_api(self, prompt: str) -> Dict[str, str]:
        """Call OpenAI API with increased token limit for comprehensive analysis."""
        try:
            response = self.ai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are a professional value investing analyst with expertise in Benjamin Graham and Warren Buffett methodologies. Provide comprehensive, detailed analysis following value investing principles. ALWAYS follow the exact format requested in the prompt."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=1500,  # Increased for comprehensive analysis
                temperature=0.2   # Lower temperature for more consistent analysis
            )
            
            ai_response = response.choices[0].message.content
            logger.info(f"Raw AI Response: {ai_response[:200]}...")  # Log first 200 chars for debugging
            
            return self._parse_ai_response(ai_response)
        except Exception as e:
            logger.error(f"OpenAI API call failed: {e}")
            return self._rule_based_analysis({})
    
    def _call_claude_api(self, prompt: str) -> Dict[str, str]:
        """Call Anthropic Claude API with enhanced settings for comprehensive analysis."""
        try:
            response = self.ai_client.messages.create(
                model="claude-3-haiku-20240307",
                max_tokens=1500,  # Increased for comprehensive analysis
                temperature=0.2,   # Lower temperature for consistent analysis
                messages=[{"role": "user", "content": prompt}]
            )
            return self._parse_ai_response(response.content[0].text)
        except Exception as e:
            logger.error(f"Claude API call failed: {e}")
            return self._rule_based_analysis({})
    
    def _call_ollama_api(self, prompt: str) -> Dict[str, str]:
        """Call local Ollama API."""
        try:
            response = requests.post('http://localhost:11434/api/generate',
                                   json={
                                       'model': 'llama2',  # or another model you have installed
                                       'prompt': prompt,
                                       'stream': False
                                   },
                                   timeout=30)
            if response.status_code == 200:
                result = response.json()
                return self._parse_ai_response(result.get('response', ''))
            else:
                raise Exception(f"Ollama API returned status {response.status_code}")
        except Exception as e:
            logger.error(f"Ollama API call failed: {e}")
            return self._rule_based_analysis({})
    
    def add_ai_analysis_to_dataframe(self, df):
        """
        Add AI analysis columns to the dataframe.
        
        Args:
            df (pd.DataFrame): Dataframe with investment candidates
            
        Returns:
            pd.DataFrame: Enhanced dataframe with AI analysis
        """
        if not self.enable_ai_analysis:
            logger.info("AI analysis is disabled. Skipping AI evaluation.")
            # Add placeholder columns for comprehensive analysis when AI is disabled
            df['AI Sentiment'] = 'Not Available'
            df['AI Recommendation'] = 'Not Available'
            df['AI Reasoning'] = 'AI analysis disabled'
            df['Value Score (1-10)'] = 'N/A'
            df['Financial Health'] = 'N/A'
            df['Business Quality'] = 'N/A'
            df['Risk Level'] = 'N/A'
            df['Valuation Assessment'] = 'N/A'
            df['Margin of Safety'] = 'N/A'
            df['Investment Thesis'] = 'N/A'
            df['Key Risks'] = 'N/A'
            df['Growth Catalysts'] = 'N/A'
            df['Target Price'] = 'N/A'
            
            # Add price prediction columns even when AI is disabled
            df['Predicted Price (30d)'] = 'N/A'
            df['Price Change %'] = 'N/A'
            df['Prediction Confidence'] = 'N/A'
            df['Prediction Method'] = 'N/A'
            
            return df
        
        logger.info("Starting comprehensive AI analysis for investment candidates...")
        
        # Initialize lists for comprehensive AI analysis
        ai_sentiments = []
        ai_recommendations = []
        ai_reasoning = []
        value_scores = []
        financial_health = []
        business_quality = []
        risk_levels = []
        valuations = []
        margin_of_safety = []
        investment_thesis = []
        key_risks = []
        catalysts = []
        target_prices = []
        
        # Price prediction columns
        predicted_prices = []
        price_change_percents = []
        prediction_confidences = []
        prediction_methods = []
        
        # Multi-period prediction columns
        price_6_months = []
        price_12_months = []
        growth_6_months = []
        growth_12_months = []
        
        total_stocks = len(df)
        
        for idx, (_, row) in enumerate(df.iterrows(), 1):
            logger.info(f"Analyzing stock {idx}/{total_stocks}: {row.get('Symbol', 'Unknown')}")
            
            # Convert row to dict for AI analysis
            row_dict = row.to_dict()
            
            # Get comprehensive AI analysis
            ai_result = self.analyze_stock_with_ai(row_dict)
            
            # Legacy fields for compatibility
            ai_sentiments.append(ai_result.get('ai_sentiment', 'N/A'))
            ai_recommendations.append(ai_result.get('ai_recommendation', 'Hold'))
            ai_reasoning.append(ai_result.get('ai_reasoning', 'N/A'))
            
            # New comprehensive analysis fields
            value_scores.append(ai_result.get('value_score', 'N/A'))
            financial_health.append(ai_result.get('financial_health', 'N/A'))
            business_quality.append(ai_result.get('business_quality', 'N/A'))
            risk_levels.append(ai_result.get('risk_level', 'N/A'))
            valuations.append(ai_result.get('valuation', 'N/A'))
            margin_of_safety.append(ai_result.get('margin_of_safety', 'N/A'))
            investment_thesis.append(ai_result.get('investment_thesis', 'N/A'))
            key_risks.append(ai_result.get('key_risks', 'N/A'))
            catalysts.append(ai_result.get('catalysts', 'N/A'))
            target_prices.append(ai_result.get('target_price', 'N/A'))
            
            # Price prediction fields
            predicted_prices.append(ai_result.get('predicted_price', 'N/A'))
            price_change_percents.append(ai_result.get('price_change_percent', 'N/A'))
            prediction_confidences.append(ai_result.get('prediction_confidence', 'N/A'))
            prediction_methods.append(ai_result.get('prediction_method', 'N/A'))
            
            # Multi-period prediction fields
            multi_period_preds = ai_result.get('multi_period_predictions', {})
            price_6_months.append(multi_period_preds.get('6_months', {}).get('predicted_price', 'N/A'))
            price_12_months.append(multi_period_preds.get('12_months', {}).get('predicted_price', 'N/A'))
            growth_6_months.append(multi_period_preds.get('6_months', {}).get('growth_percent', 'N/A'))
            growth_12_months.append(multi_period_preds.get('12_months', {}).get('growth_percent', 'N/A'))
            
            # Small delay to avoid API rate limits
            if idx < total_stocks:
                time.sleep(1)  # 1 second delay between API calls
        
        # Add comprehensive AI analysis columns
        df['AI Sentiment'] = ai_sentiments
        df['AI Recommendation'] = ai_recommendations  
        df['AI Reasoning'] = ai_reasoning
        
        # Value Investing Analysis Columns
        df['Value Score (1-10)'] = value_scores
        df['Financial Health'] = financial_health
        df['Business Quality'] = business_quality
        df['Risk Level'] = risk_levels
        df['Valuation Assessment'] = valuations
        df['Margin of Safety'] = margin_of_safety
        df['Investment Thesis'] = investment_thesis
        df['Key Risks'] = key_risks
        df['Growth Catalysts'] = catalysts
        df['Target Price'] = target_prices
        
        # Price Prediction Columns
        df['Predicted Price (30d)'] = predicted_prices
        df['Price Change %'] = price_change_percents
        df['Prediction Confidence'] = prediction_confidences
        df['Prediction Method'] = prediction_methods
        
        # Multi-Period Prediction Columns
        df['Price Target (6M)'] = price_6_months
        df['Price Target (12M)'] = price_12_months
        df['Growth 6M (%)'] = growth_6_months
        df['Growth 12M (%)'] = growth_12_months
        
        logger.info("Comprehensive AI analysis completed for all stocks")
        
        # Log analysis distribution
        sentiment_counts = pd.Series(ai_sentiments).value_counts()
        valid_scores = [float(x) for x in value_scores if str(x).replace('.','').replace('-','').isdigit()]
        if valid_scores:
            value_score_avg = sum(valid_scores) / len(valid_scores)
            logger.info(f"Average Value Score: {value_score_avg:.1f}/10")
        logger.info(f"AI Sentiment Distribution: {dict(sentiment_counts)}")
        
        return df
    
    def save_detailed_analysis(self, df):
        """
        Save the detailed analysis to Excel with enhanced formatting.
        
        Args:
            df (pd.DataFrame): Processed data with additional analysis
        """
        # Ensure output directory exists
        os.makedirs(os.path.dirname(self.output_file), exist_ok=True)
        
        # Sort by available score column (highest first)
        if 'Investment Readiness Score' in df.columns:
            df_sorted = df.sort_values('Investment Readiness Score', ascending=False)
        elif 'Value Score (1-10)' in df.columns:
            df_sorted = df.sort_values('Value Score (1-10)', ascending=False)
        else:
            # Fallback to Symbol sorting if no score columns available
            df_sorted = df.sort_values('Symbol', ascending=True)
        
        # Save to Excel
        df_sorted.to_excel(self.output_file, index=False)
        logger.info(f"Saved detailed analysis to: {self.output_file}")
        
        # Apply formatting
        self._apply_excel_formatting()
        
    def _apply_excel_formatting(self):
        """Apply professional formatting to the Excel file."""
        wb = load_workbook(self.output_file)
        ws = wb.active
        
        # Define colors for readiness categories
        category_colors = {
            'Excellent': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),      # Green
            'Very Good': PatternFill(start_color='92D050', end_color='92D050', fill_type='solid'),     # Light Green
            'Good': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),          # Yellow
            'Fair': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),          # Orange
            'Needs Review': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')   # Red
        }
        
        # Define colors for AI sentiment
        sentiment_colors = {
            'Good': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),          # Green
            'Bad': PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'),           # Orange  
            'Worst': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),         # Red
            'Neutral': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),       # Gray
            'Not Available': PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid'), # Light Gray
            'Error': PatternFill(start_color='800080', end_color='800080', fill_type='solid')          # Purple
        }
        
        # Define colors for AI recommendations  
        ai_rec_colors = {
            'Strong Buy': PatternFill(start_color='006100', end_color='006100', fill_type='solid'),    # Dark Green
            'Buy': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),          # Green
            'Hold': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),         # Yellow
            'Avoid': PatternFill(start_color='C00000', end_color='C00000', fill_type='solid'),        # Dark Red
            'Not Available': PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid') # Light Gray
        }
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        
        # Find relevant columns
        readiness_score_col = None
        readiness_category_col = None
        recommendation_col = None
        ai_sentiment_col = None
        ai_recommendation_col = None
        ai_reasoning_col = None
        
        for i, header in enumerate(headers):
            if header == 'Investment Readiness Score':
                readiness_score_col = i + 1
            elif header == 'Readiness Category':
                readiness_category_col = i + 1
            elif header == 'Investment Recommendation':
                recommendation_col = i + 1
            elif header == 'AI Sentiment':
                ai_sentiment_col = i + 1
            elif header == 'AI Recommendation':
                ai_recommendation_col = i + 1
            elif header == 'AI Reasoning':
                ai_reasoning_col = i + 1
        
        # Apply formatting to data rows
        for row_idx in range(2, ws.max_row + 1):
            # Color code readiness category
            if readiness_category_col:
                category_cell = ws.cell(row=row_idx, column=readiness_category_col)
                category = category_cell.value
                if category in category_colors:
                    category_cell.fill = category_colors[category]
                    if category in ['Excellent', 'Needs Review']:
                        category_cell.font = Font(bold=True, color='FFFFFF')
                    else:
                        category_cell.font = Font(bold=True)
            
            # Bold and color the readiness score
            if readiness_score_col:
                score_cell = ws.cell(row=row_idx, column=readiness_score_col)
                score_value = score_cell.value
                if score_value is not None:
                    try:
                        score = float(score_value)
                        if score >= 80:
                            score_cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
                            score_cell.font = Font(bold=True, color='FFFFFF')
                        elif score >= 70:
                            score_cell.fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                            score_cell.font = Font(bold=True)
                        elif score < 50:
                            score_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            score_cell.font = Font(bold=True, color='FFFFFF')
                        else:
                            score_cell.font = Font(bold=True)
                    except (ValueError, TypeError):
                        pass
            
            # Highlight investment recommendations
            if recommendation_col:
                rec_cell = ws.cell(row=row_idx, column=recommendation_col)
                if rec_cell.value == 'Strong Buy':
                    rec_cell.fill = PatternFill(start_color='006100', end_color='006100', fill_type='solid')
                    rec_cell.font = Font(bold=True, color='FFFFFF')
                elif rec_cell.value == 'Buy':
                    rec_cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
                    rec_cell.font = Font(bold=True, color='FFFFFF')
            
            # Color code AI sentiment
            if ai_sentiment_col:
                sentiment_cell = ws.cell(row=row_idx, column=ai_sentiment_col)
                sentiment = sentiment_cell.value
                if sentiment in sentiment_colors:
                    sentiment_cell.fill = sentiment_colors[sentiment]
                    if sentiment in ['Good', 'Worst', 'Error']:
                        sentiment_cell.font = Font(bold=True, color='FFFFFF')
                    else:
                        sentiment_cell.font = Font(bold=True)
            
            # Color code AI recommendations
            if ai_recommendation_col:
                ai_rec_cell = ws.cell(row=row_idx, column=ai_recommendation_col)
                ai_rec = ai_rec_cell.value
                if ai_rec in ai_rec_colors:
                    ai_rec_cell.fill = ai_rec_colors[ai_rec]
                    if ai_rec in ['Strong Buy', 'Buy', 'Avoid']:
                        ai_rec_cell.font = Font(bold=True, color='FFFFFF')
                    else:
                        ai_rec_cell.font = Font(bold=True)
            
            # Style AI reasoning column for better readability
            if ai_reasoning_col:
                reasoning_cell = ws.cell(row=row_idx, column=ai_reasoning_col)
                if reasoning_cell.value and reasoning_cell.value != 'Not Available':
                    reasoning_cell.alignment = Alignment(wrap_text=True, vertical='top')
                    reasoning_cell.font = Font(size=9)  # Smaller font for reasoning
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        
        wb.save(self.output_file)
        logger.info("Applied professional formatting to Excel file")
    
    def _calculate_avg_value_score(self, df):
        """Calculate average AI value score from the dataframe."""
        if 'Value Score (1-10)' not in df.columns:
            return 'N/A'
        
        value_scores = df['Value Score (1-10)']
        valid_scores = []
        
        for score in value_scores:
            if score != 'N/A' and str(score).replace('.','').replace('-','').isdigit():
                try:
                    valid_scores.append(float(score))
                except:
                    continue
        
        if valid_scores:
            return round(sum(valid_scores) / len(valid_scores), 1)
        return 'N/A'
    
    def run_analysis(self):
        """
        Execute the complete detailed analysis workflow.
        
        Returns:
            dict: Summary statistics of the analysis
        """
        try:
            logger.info("Starting comprehensive analysis of all stocks...")
            
            # Load data
            df = self.load_screener_data()
            
            if len(df) == 0:
                logger.warning("No stocks found in the data!")
                return {'candidates_found': 0}
            
            logger.info(f"Processing {len(df)} stocks for comprehensive analysis...")
            
            # Calculate additional analysis for all stocks
            enhanced_df = self.calculate_investment_readiness(df)
            
            # Add AI-powered comprehensive analysis (or placeholder columns if disabled)
            enhanced_df = self.add_ai_analysis_to_dataframe(enhanced_df)
            if self.enable_ai_analysis:
                logger.info("Comprehensive AI analysis integration completed")
            else:
                logger.info("AI placeholder columns added")
            
            # Save results
            self.save_detailed_analysis(enhanced_df)
            
            # Generate summary
            summary = {
                'total_stocks_analyzed': len(df),
                'candidates_found': len(enhanced_df),
                'strong_buy_count': len(enhanced_df[enhanced_df['Investment Recommendation'] == 'Strong Buy']),
                'buy_count': len(enhanced_df[enhanced_df['Investment Recommendation'] == 'Buy']),
                'hold_count': len(enhanced_df[enhanced_df['Investment Recommendation'].str.contains('Hold', na=False)]),
                'avg_readiness_score': round(enhanced_df['Investment Readiness Score'].mean(), 1) if 'Investment Readiness Score' in enhanced_df.columns else 0,
                'excellent_candidates': len(enhanced_df[enhanced_df['Readiness Category'] == 'Excellent']) if 'Readiness Category' in enhanced_df.columns else 0,
                'avg_ai_value_score': self._calculate_avg_value_score(enhanced_df),
                'output_file': self.output_file
            }
            
            logger.info("=== COMPREHENSIVE ANALYSIS COMPLETE ===")
            logger.info(f"Total stocks analyzed: {summary['total_stocks_analyzed']}")
            logger.info(f"Stocks processed: {summary['candidates_found']}")
            logger.info(f"  - Strong Buy: {summary['strong_buy_count']}")
            logger.info(f"  - Buy: {summary['buy_count']}")
            logger.info(f"  - Hold/Others: {summary['hold_count']}")
            logger.info(f"Average Readiness Score: {summary['avg_readiness_score']}/100")
            logger.info(f"Average AI Value Score: {summary['avg_ai_value_score']}/10")
            logger.info(f"Excellent candidates: {summary['excellent_candidates']}")
            logger.info(f"Results saved to: {summary['output_file']}")
            
            return summary
            
        except Exception as e:
            logger.error(f"Analysis failed: {str(e)}")
            raise


def print_usage():
    """Print usage information for the detailed analysis tool."""
    print("""
🔍 Stock Screener - Detailed Analysis Tool

USAGE:
    python detailed_analysis.py [input_file] [output_file] [options]

OPTIONS:
    --no-ai                 Disable AI analysis (use rule-based only)
    --ai-provider=PROVIDER  Specify AI provider (auto, openai, gemini, claude, ollama)
    --openai               Use OpenAI GPT (shorthand for --ai-provider=openai)
    --gemini               Use Google Gemini (shorthand for --ai-provider=gemini) 
    --claude               Use Anthropic Claude (shorthand for --ai-provider=claude)
    --ollama               Use local Ollama (shorthand for --ai-provider=ollama)
    --help, -h             Show this help message

EXAMPLES:
    # Auto-detect best available AI provider
    python detailed_analysis.py
    
    # Force specific AI provider
    python detailed_analysis.py --gemini
    python detailed_analysis.py --ai-provider=openai
    
    # Use custom input/output files
    python detailed_analysis.py data/my_input.xlsx data/my_output.xlsx --claude
    
    # Disable AI analysis completely
    python detailed_analysis.py --no-ai

AI PROVIDERS:
    • auto     - Auto-detect best available provider (default)
    • openai   - OpenAI GPT-4o-mini (requires OPENAI_API_KEY)
    • gemini   - Google Gemini 1.5 Flash (requires GOOGLE_API_KEY - FREE tier available)
    • claude   - Anthropic Claude Haiku (requires ANTHROPIC_API_KEY)
    • ollama   - Local Ollama server (no API key needed)

DEFAULT PATHS:
    Input:  data/output/value_analysis.xlsx
    Output: data/output/detailed_analysis.xlsx
""")


def main():
    """Main function to run the detailed analysis."""
    # Parse command line arguments
    args = [arg for arg in sys.argv[1:] if not arg.startswith('--')]
    input_file = args[0] if len(args) > 0 else None
    output_file = args[1] if len(args) > 1 else None
    
    # Try to load AI settings from config first
    config_file = 'config/screener_config.properties'
    enable_ai = True
    preferred_provider = 'auto'
    
    if os.path.exists(config_file):
        try:
            import configparser
            config = configparser.ConfigParser()
            config.read(config_file)
            enable_ai = config.getboolean('DEFAULT', 'ai_analysis_enabled', fallback=True)
            preferred_provider = config.get('DEFAULT', 'ai_provider', fallback='auto')
            logger.info(f"AI settings loaded from config: enabled={enable_ai}, provider={preferred_provider}")
        except Exception as e:
            logger.warning(f"Failed to load config file: {e}")
    
    # Command line arguments can override config settings
    if '--no-ai' in sys.argv:
        enable_ai = False
        logger.info("AI analysis disabled via command line flag")
    
    # Check for AI provider preference (command line overrides config)
    for arg in sys.argv:
        if arg.startswith('--ai-provider='):
            preferred_provider = arg.split('=')[1].lower()
            if preferred_provider not in ['auto', 'openai', 'gemini', 'claude', 'ollama']:
                print(f"❌ Invalid AI provider: {preferred_provider}")
                print("Valid options: auto, openai, gemini, claude, ollama")
                sys.exit(1)
            break
        elif arg in ['--openai', '--gpt']:
            preferred_provider = 'openai'
        elif arg in ['--gemini', '--google']:
            preferred_provider = 'gemini'
        elif arg in ['--claude', '--anthropic']:
            preferred_provider = 'claude'
        elif arg == '--ollama':
            preferred_provider = 'ollama'
        elif arg in ['--help', '-h']:
            print_usage()
            sys.exit(0)
    
    # Create and run analyzer
    analyzer = DetailedAnalyzer(input_file, output_file, enable_ai_analysis=enable_ai, 
                               preferred_ai_provider=preferred_provider, config_file=config_file)
    
    # Display AI analysis status
    if analyzer.enable_ai_analysis:
        print("🤖 AI Analysis: ENABLED")
        print(f"   - Preferred Provider: {preferred_provider}")
        print(f"   - Active Provider: {analyzer.ai_provider if hasattr(analyzer, 'ai_provider') else 'Initializing...'}")
        print("   - LLM-powered ratio analysis")  
        print("   - News sentiment evaluation")
        print("   - Independent AI recommendations")
    else:
        print("📊 AI Analysis: DISABLED")
        print("   - Standard quantitative analysis only")
        if not HAS_AI:
            print("   - Install AI provider packages to enable AI features")
            print("     • Google Gemini (free): pip install google-generativeai")
            print("     • OpenAI: pip install openai") 
            print("     • Anthropic Claude: pip install anthropic")
            print("     • Or run Ollama locally")
        elif not any([os.getenv(key) for key in ['GOOGLE_API_KEY', 'OPENAI_API_KEY', 'ANTHROPIC_API_KEY']]):
            print("   - Set API key environment variable to enable AI")
    
    try:
        summary = analyzer.run_analysis()
        
        if summary['candidates_found'] > 0:
            print("\n🎯 COMPREHENSIVE ANALYSIS COMPLETE!")
            print(f"📊 Analyzed {summary['candidates_found']} stocks")
            print(f"📈 Average readiness: {summary['avg_readiness_score']}/100")
            if summary['avg_ai_value_score'] != 'N/A':
                print(f"🧠 Average AI value score: {summary['avg_ai_value_score']}/10")
            print(f"⭐ Excellent candidates: {summary['excellent_candidates']}")
            if summary['strong_buy_count'] > 0 or summary['buy_count'] > 0:
                print(f"🚀 Strong Buy: {summary['strong_buy_count']}, Buy: {summary['buy_count']}")
            print(f"📄 Detailed report: {summary['output_file']}")
            
            if analyzer.enable_ai_analysis:
                print("\n🤖 Enhanced with comprehensive AI analysis:")
                print("   - Value investing scorecard (Graham & Buffett principles)")
                print("   - Financial health & business quality assessment") 
                print("   - Risk analysis & margin of safety calculation")
                print("   - Investment thesis & growth catalysts")
                print("   - Target price estimates & valuation analysis")
        else:
            print("\n⚠️  No stocks found in analysis!")
            print("Run the main screener first to generate data.")
            
    except Exception as e:
        print(f"\n❌ Analysis failed: {str(e)}")
        if "API key" in str(e).lower():
            print("💡 Tip: Set appropriate API key environment variable for AI analysis")
            print("   - Google Gemini: GOOGLE_API_KEY")
            print("   - OpenAI: OPENAI_API_KEY") 
            print("   - Anthropic Claude: ANTHROPIC_API_KEY")
        sys.exit(1)


if __name__ == "__main__":
    main()
