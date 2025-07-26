import pandas as pd
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import os


class ExcelExporter:
    def __init__(self, output_path='data/output/value_analysis.xlsx'):
        self.output_path = output_path

        # Define fill styles for conditional formatting
        self.pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Green
        self.fail_fill = PatternFill(start_color='F2DCDB', end_color='F2DCDB', fill_type='solid')  # Red
        self.neutral_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Yellow

    def write_data(self, results: list[dict]):
        try:
            # Convert results into a DataFrame
            df = pd.DataFrame(results)
            logging.info(f"Initial DataFrame created with {len(df)} rows.")

            # Replace NaN in object-type columns with "N/A"
            df.fillna("N/A", inplace=True)

            '''for col in df.select_dtypes(include='object').columns:
                df[col] = df[col].fillna("N/A")'''

            # Identify metric columns (exclude non-metric ones)
            non_metric_cols = ['Symbol', 'Investment Recommendation']
            metric_columns = [col for col in df.columns if col not in non_metric_cols and not col.endswith(" Pass")]

            # Set recommendation to 'Insufficient data' if all metrics are 'N/A' or NaN
            for index, row in df.iterrows():
                all_na = all(row[col] in ['N/A', None, float('nan')] for col in metric_columns)
                if all_na:
                    df.at[index, 'Investment Recommendation'] = 'Insufficient data'
                    logging.info(f"Row {index} marked as 'Insufficient data' due to missing metrics.")

            # Save DataFrame to Excel
            os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
            df.to_excel(self.output_path, index=False)
            logging.info(f"Excel file saved at {self.output_path}")

            # Load workbook for styling
            wb = load_workbook(self.output_path)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            # Identify *_Pass columns and their base metric
            pass_columns = [col for col in headers if col.endswith("Pass") and col != "Investment Recommendation"]
            base_metrics = [col.replace(" Pass", "") for col in pass_columns]

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                row_map = {headers[i]: cell for i, cell in enumerate(row)}

                # Apply fill color based on pass/fail columns
                for metric in base_metrics:
                    metric_cell = row_map.get(metric)
                    pass_cell = row_map.get(f"{metric} Pass")
                    if metric_cell and pass_cell:
                        if pass_cell.value is True:
                            metric_cell.fill = self.pass_fill
                        elif pass_cell.value is False:
                            metric_cell.fill = self.fail_fill

                # Apply styling to Investment Recommendation cell
                rec_cell = row_map.get("Investment Recommendation")
                if rec_cell:
                    if rec_cell.value == "Strong Buy":
                        rec_cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
                        rec_cell.font = Font(bold=True)
                    elif rec_cell.value == "Hold":
                        rec_cell.fill = self.neutral_fill
                    elif rec_cell.value == "Avoid":
                        rec_cell.fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
                        rec_cell.font = Font(bold=True)
                    elif rec_cell.value == "Insufficient data":
                        rec_cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                        rec_cell.font = Font(bold=True, italic=True)

            # Remove *_Pass columns from the sheet
            for col in reversed(pass_columns):  # Reverse to avoid index shifting
                col_index = headers.index(col) + 1
                ws.delete_cols(col_index)
                headers.pop(col_index - 1)
                logging.info(f"Dropped column: {col}")

            wb.save(self.output_path)
            logging.info("Styling and cleanup complete. File saved.")

        except Exception as e:
            logging.error(f"An error occurred during Excel export: {e}", exc_info=True)
