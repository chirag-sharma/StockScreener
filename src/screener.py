# src/screener.py

import yfinance as yf
import pandas as pd
from configparser import ConfigParser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from utils.ticker_loader import load_tickers

def load_symbols_from_config():
    """
    Loads stock symbols for a given sector from the configuration file.

    Returns:
        dict: Mapping of company names to ticker symbols.

    Raises:
        ValueError: If the sector is not specified in the config.
    """
    config = ConfigParser()
    config.read('../config/screener_config.properties')
    sector = config.get('DEFAULT', 'sector', fallback=None)
    if not sector:
        raise ValueError("Sector not specified in config.")
    symbols = load_tickers(sector)
    return symbols

def analyze_stock(symbol):
    """
    Analyzes a stock based on key financial metrics.

    Args:
        symbol (str): Ticker symbol of the stock.

    Returns:
        dict: Analysis results including metrics and pass/fail status.
    """
    try:
        stock = yf.Ticker(symbol)
        info = stock.info

        # Extract relevant financial metrics
        analysis = {
            'Symbol': symbol,
            'PE Ratio': info.get('trailingPE', None),
            'Debt/Equity': info.get('debtToEquity', None),
            'ROE': info.get('returnOnEquity', None) * 100 if info.get('returnOnEquity') else None,
            'Current Ratio': info.get('currentRatio', None),
            'Price to Book': info.get('priceToBook', None),
            'Promoter Holding': info.get('heldPercentInsiders', None) * 100 if info.get('heldPercentInsiders') else None
        }

        # Define criteria for value analysis
        criteria = {
            'PE Ratio': analysis['PE Ratio'] is not None and analysis['PE Ratio'] < 20,
            'Debt/Equity': analysis['Debt/Equity'] is not None and analysis['Debt/Equity'] < 1,
            'ROE': analysis['ROE'] is not None and analysis['ROE'] > 15,
            'Current Ratio': analysis['Current Ratio'] is not None and analysis['Current Ratio'] > 1.5,
            'Price to Book': analysis['Price to Book'] is not None and analysis['Price to Book'] < 3,
            'Promoter Holding': analysis['Promoter Holding'] is not None and analysis['Promoter Holding'] > 50
        }

        # Add pass/fail status for each metric
        for key in criteria:
            analysis[key + ' Pass'] = criteria[key]

        return analysis

    except Exception as e:
        print(f"Error analyzing {symbol}: {e}")
        return None

def write_to_excel(results, filename='value_analysis.xlsx'):
    """
    Writes analysis results to an Excel file and highlights failed criteria.

    Args:
        results (list): List of analysis dictionaries.
        filename (str): Output Excel file name.
    """
    df = pd.DataFrame(results)

    # Save results to Excel
    output_path = f'../data/output/{filename}'
    df.to_excel(output_path, index=False)

    # Load workbook and worksheet for formatting
    wb = load_workbook(output_path)
    ws = wb.active

    # Define red fill for failed criteria
    red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')

    # Highlight cells where criteria failed
    for row in range(2, ws.max_row + 1):  # Skip header row
        for col_idx, field in enumerate(['PE Ratio', 'Debt/Equity', 'ROE', 'Current Ratio', 'Price to Book', 'Promoter Holding'], start=2):
            pass_col = col_idx + 6  # Offset to corresponding "Pass" column
            if ws.cell(row=row, column=pass_col).value is False:
                ws.cell(row=row, column=col_idx).fill = red_fill

    wb.save(output_path)
    print(f"\n✅ Analysis written to {output_path}")

def main():
    """
    Main execution flow:
    - Loads symbols for the configured sector.
    - Performs analysis for each symbol.
    - Writes results to Excel.
    """
    symbols_dict = load_symbols_from_config()
    results = []

    for name, symbol in symbols_dict.items():
        print(f"Analyzing {name} ({symbol})...")
        data = analyze_stock(symbol)
        if data:
            results.append(data)

    write_to_excel(results)

if __name__ == "__main__":
    main()